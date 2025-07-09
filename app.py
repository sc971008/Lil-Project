from flask import Flask, render_template, request, send_file
import os
import re
import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime

app = Flask(__name__)
UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def parse_event_time(event_time_str):
    # 示例格式：Tue Jul 08 2025 18:51:50 GMT-0400 (EDT)
    m = re.search(r"(\w{3} \w{3} \d{1,2} \d{4} \d{2}:\d{2}:\d{2})\s*GMT([+-]\d{4})", str(event_time_str))
    if m:
        dt_str = f"{m.group(1)} {m.group(2)}"
        try:
            dt_obj = datetime.strptime(dt_str, "%a %b %d %Y %H:%M:%S %z")
            return dt_obj.strftime("%m/%d/%y %H:%M")
        except Exception as e:
            return ""
    return ""

def extract_from_excel(filepath):
    df = pd.read_csv(filepath) if filepath.endswith('.csv') else pd.read_excel(filepath)

    # 确保必要的列存在
    required_columns = {"Gov Agency", "Entry Number", "Line#", "Event Time"}
    if not required_columns.issubset(df.columns):
        raise ValueError(f"Excel/CSV 缺少必要列：{required_columns - set(df.columns)}")

    # 清洗并定位最新一段以 "Date/Time" 开头的段落
    pga_indices = df[df["Event Time"].astype(str).str.startswith("Date/Time")].index.tolist()
    if not pga_indices:
        print("没有找到任何以 'Date/Time' 开头的记录")
        return []

    # 最新的一段（最上面那段）的位置
    latest_index = pga_indices[0]

    # 截取从该行开始的所有记录
    df_latest = df.iloc[latest_index:].copy()

    records = []

    for _, row in df_latest.iterrows():
        if str(row.get("Gov Agency", "")).strip() != "CPS":
            continue  # 只处理 Gov Agency 是 CPS 的记录

        entry_raw = str(row.get("Entry Number", "")).strip()
        line_no = str(row.get("Line#", "")).strip()
        event_time_raw = row.get("Event Time", "")
        event_time = parse_event_time(event_time_raw)

        formatted_entry = (
            f"NVB-{entry_raw[:7]}-{entry_raw[-1]}"
            if len(entry_raw) >= 8 else entry_raw
        )

        records.append({
            "entry": formatted_entry,
            "status": "CPSC_check",
            "event_time": event_time,
            "timezone": "America/New_York",
            "line": line_no
        })
        return records

def generate_excel(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CPSC Check"

    headers = ["Entry Number", "Status", "Event Time", "Time Zone", "Line"]
    ws.append(headers)

    for item in data:
        ws.append([
            item["entry"],
            item["status"],
            item["event_time"],
            item["timezone"],
            item["line"]
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    output = BytesIO()
    today_str = datetime.now().strftime("%Y%m%d")
    filename = f"CPSC_Check_Results_{today_str}.xlsx"
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        uploaded_files = request.files.getlist("excels")
        all_records = []
        for file in uploaded_files:
            if file and (file.filename.endswith('.xlsx') or file.filename.endswith('.csv')):
                filepath = os.path.join(UPLOAD_FOLDER, file.filename)
                file.save(filepath)
                all_records += extract_from_excel(filepath)

        # 清理上传目录
        for file in os.listdir(UPLOAD_FOLDER):
            try:
                os.remove(os.path.join(UPLOAD_FOLDER, file))
            except Exception as e:
                print(f"Failed to delete {file}: {e}")

        if all_records:
            return generate_excel(all_records)
    return render_template("index.html")

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
